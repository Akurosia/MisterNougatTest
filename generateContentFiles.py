#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# coding: utf8
import copy
import os
from operator import itemgetter
import traceback
import re
import errno
import openpyxl
import yaml
from yaml.loader import SafeLoader
import math
import natsort
from collections import OrderedDict
import logging
from openpyxl import load_workbook
from io import BytesIO
import requests

import sys
from ffxiv_aku import *

enemy = {
    "title": "",
    "title_en": "",
    "id": "",
    "hp": {
        "min": 100,
        "max": 0
    },
    "attacks": [{
        "title": "",
        "title_id": "",
        "title_en": "",
        "attack_in_use": "",
        "disable": "",
        "type": "",
        "damage_type": "",
        "damage": {
            "min": 100,
            "max": 0
        },
        "phases": [{"phase": "", }],
        "roles": [{"role": "", }],
        "tags": [{"tag": "", }],
        "notes": [{"note": "", }],
    }],
}
example_sequence = {
    "sequence": [{
        "phase": "09",
        "name": "phase_name",
        "alerts": [{
            "alert": "Die folgenden Angriffe haben sind entweder unbekannt oder haben keine klare Herkunft",
        }],
        "mechanics": [{
            "title": "sequence-mechanic-01",
            "notes": [{
                "note": "sequence-mechanic-note-01",
            }],
        }],
        "attacks": [{
            "attack": "sequence-attack-01",
        }],
        "images": [{
            "url": "/assets/img/test.jpg",
            "alt": "/assets/img/test.jpg",
            "height": "250px",
        }],
        "videos": [{
            "url": "https&#58;//ffxivguide.akurosia.de/upload/test.mp4",
        }],
    }]
}
example_add_sequence = {
    "sequence": [{"phase": "09", }]
}


class CustomFormatter(logging.Formatter):
    grey = "\x1b[38;20m"
    yellow = "\x1b[33;20m"
    red = "\x1b[31;20m"
    bold_red = "\x1b[31;1m"
    reset = "\x1b[0m"
    format = "%(asctime)s - %(name)s - %(levelname)s - %(message)s (%(filename)s:%(lineno)d)"
    FORMATS = {
        logging.DEBUG: wrap_in_color_green(format),
        logging.INFO: wrap_in_color_blue(format),
        logging.WARNING: wrap_in_color_yellow(format),
        logging.ERROR: wrap_in_color_red(format),
        logging.CRITICAL: bold_red + format + reset
    }

    def format(self, record):
        log_fmt = self.FORMATS.get(record.levelno)
        formatter = logging.Formatter(log_fmt)
        return formatter.format(record)


logger = logging.getLogger("My_app")
logger.setLevel(logging.CRITICAL)
ch = logging.StreamHandler()
ch.setLevel(logging.CRITICAL)
ch.setFormatter(CustomFormatter())
logger.addHandler(ch)

disable_green_print = True
disable_yellow_print = True
disable_blue_print = True
disable_red_print = True

storeFilesInTmp(True)
logdata = get_any_Logdata()
patchversions = get_any_Versiondata()
logdata_lower = dict((k.lower(), v) for k, v in logdata.items())
action = loadDataTheQuickestWay("action_all.json", translate=True)
territorytype = loadDataTheQuickestWay("territorytype_all.json", translate=True)
contentfindercondition = loadDataTheQuickestWay("contentfindercondition_all.json", translate=True)
contentfinderconditionX = loadDataTheQuickestWay("ContentFinderCondition.de.json")
contentmembertype = loadDataTheQuickestWay("ContentMemberType.json")
quests = loadDataTheQuickestWay("Quest.de.json")
levels = loadDataTheQuickestWay("Level.json")
maps = loadDataTheQuickestWay("Map.json")
placename = loadDataTheQuickestWay("placename_all.json", translate=True)
bnpcname = loadDataTheQuickestWay("bnpcname_all.json", translate=True)
eobjname = loadDataTheQuickestWay("eobjname_all.json", translate=True)
status = loadDataTheQuickestWay("status_all.json", translate=True)
enpcresident = loadDataTheQuickestWay("enpcresident_all.json", translate=True)
mounts = loadDataTheQuickestWay("mount_all.json", translate=True)
minions = loadDataTheQuickestWay("companion_all.json", translate=True)
orchestrions = loadDataTheQuickestWay("orchestrion_all.json", translate=True)
ttcards = loadDataTheQuickestWay("tripletriadcard_all.json", translate=True)
LANGUAGES = ["de", "en", "fr", "ja", "cn", "ko"]
XLSXELEMENTS = ["exclude", "date", "sortid", "title", "categories", "slug", "image", "patchNumber", "patchName", "difficulty", "plvl", "plvl_sync", "ilvl", "ilvl_sync", "quest_id", "gearset_loot", "tt_card1", "tt_card2", "orchestrion", "orchestrion2", "orchestrion3", "orchestrion4", "orchestrion5", "orchestrion_material1", "orchestrion_material2", "orchestrion_material3", "mtqvid1", "mtqvid2", "mrhvid1", "mrhvid2", "mount1", "mount2", "minion1", "minion2", "minion3", "instanceType", "mapid", "bosse", "adds", "mechanics", "tags", "teamcraftlink", "garlandtoolslink", "gamerescapelink", "done"]
UNKNOWNTITLE = {'de': 'Unbekannte Herkunft', 'en': 'Unknown Source', 'fr': 'Unknown Source', 'ja': 'Unknown Source', 'cn': 'Unknown Source', 'ko': 'Unknown Source'}


def load_workbook_from_url(url):
    file = requests.get(url)
    return load_workbook(filename = BytesIO(file.content))


def read_xlsx_file():
    # open file, get sheet, last row and last coulmn
    wb = load_workbook_from_url('https://raw.githubusercontent.com/Akurosia/DevFFXIVPocketGuide/master/guide_ffxiv.xlsx')
    sheet = wb['Tabelle1']
    max_row = sheet.max_row
    max_column = sheet.max_column
    return sheet, max_row, max_column


def getPrevAndNextContentOrder(sheet, elements, max_row):
    entry = {}
    for i in range(1, max_row + 1):
        instanceType = str(sheet.cell(row=int(i), column=int(elements.index('instanceType')) + 1).value).replace("None", "")
        if not entry.get(instanceType, None):
            entry[instanceType] = {}
        sortID = str(sheet.cell(row=int(i), column=int(3)).value).replace("None", "")
        addon = str(sheet.cell(row=int(i), column=int(5)).value).replace("None", "")
        slug = str(sheet.cell(row=int(i), column=int(6)).value).replace("None", "")
        entry[instanceType][sortID] = "/" + addon + "/" + slug
    return OrderedDict(natsort.natsorted(entry.items()))


def get_data_from_xlsx(sheet, max_column, i, elements):
    entry = {}
    # for every column in row add all elements into a dict:
    # max_column will ignore last column due to how range is working
    for j in range(1, max_column + 1):
        entry[elements[j - 1]] = str(sheet.cell(row=int(i), column=int(j)).value).replace("None", "")
    return entry


def clean_entries_from_single_quotes(entry):
    for key, value in entry.items():
        if value.startswith("'"):
            entry[key] = value[1:]
        if value.endswith("'"):
            entry[key] = value[:-1]
    return entry

#getlevel
def workOnQuests(entry, quest_id):
    global quests
    if quest_id == "":
        entry['quest'] = ""
        entry['quest_location'] = ""
        entry['quest_npc'] = ""
        return entry
    quest = quests[quest_id]
    entry['quest'] = quest['Name'].replace(" ", "").replace(" ", "")
    entry['quest_npc'] = quest['Issuer']['Start']
    try:
        level_data = getLevel(quest['Issuer']['Location'])
        entry['quest_location'] = f'{level_data["placename"]} ({level_data["x"]}, {level_data["y"]})'
    except KeyError:
        entry['quest_location'] = ""
        print_color_red(f"[workOnQuests] Error on loading: {quest['Issuer']['Location']} ({quest_id})")
    return entry


def uglyContentNameFix(name, instanceType=None, difficulty=None):
    if difficulty == "Fatal" and instanceType == "ultimate" and "fatal" not in name.lower():
        name = f"{name} (fatal)"
    elif difficulty == "Episch" and instanceType == "raid" and "episch" not in name.lower():
        name = f"{name} (episch)"
    elif difficulty == "Episch" and instanceType == "feldexkursion" and "episch" not in name.lower():
        name = f"{name} (episch)"
    elif difficulty == "Schwer" and instanceType == "dungeon" and "schwer" not in name.lower():
        name = f"{name} (schwer)"
    # handle stupid edge cases for primals
    elif name in ["Königliche Konfrontation", "Jagd auf Rathalos"] and difficulty.lower() != "normal":
        name = f"{name} ({difficulty.lower()})"
    elif name in ["Memoria Misera"] and difficulty.lower() != "normal":
        name = f"{name} ({difficulty.lower()})"
    elif name in ["Krieger des Lichts"] and difficulty.lower() == "extrem":
        name = f"{name} ({difficulty.lower()})"
    # handle edge cases PvP
    elif "(Flechtenhain)" in name:
        name = name.replace("(Flechtenhain)", "")
    elif "(Kampfplatz)" in name:
        name = name.replace("(Kampfplatz)", "")
    # placeholder
    elif name == "":
        return ""
    # make sure brackets are always lowercase
    name = name.replace("(Fatal)", "(fatal)").replace("(Episch)", "(episch)").replace("(Schwer)", "(schwer)").replace("(Extrem)", "(extrem)")
    return name


def getContentName(name, lang="en", difficulty=None, instanceType=None):
    name = uglyContentNameFix(name, instanceType, difficulty)
    try:
        for key, content in contentfindercondition.items():
            if "memoria" in content["Name_de"].lower().strip() and "memoria" in name.lower().strip():
                return content[f"Name_{lang}"]
            if content["Name_de"].lower().strip() == name.lower().strip():
                return content[f"Name_{lang}"]
        for key, place in placename.items():
            if place["Name_de"].lower().strip() == name.lower().strip():
                return place[f"Name_{lang}"]
    except KeyError:
        pass
    if name not in ['title']:
        print_color_red("Could not translate: " + name)
    return ""


def getEntriesForRouletts(entry):
    global contentfinderconditionX
    for key, value in contentfinderconditionX.items():
        if value['Name'] == getContentName(entry["title"], "de", entry["difficulty"], entry["instanceType"]):
            entry['type'] = value['ContentType'].lower()
            entry['mapid'] = value['TerritoryType']
            entry['allianceraid'] = value['AllianceRoulette']
            entry['frontier'] = value['FeastTeamRoulette']
            entry['expert'] = value['ExpertRoulette']
            entry['guildhest'] = value['GuildHestRoulette']
            entry['level50_60_70'] = value['Level50/60/70Roulette']
            entry['level80'] = value['Level80Roulette']
            entry['leveling'] = value['LevelingRoulette']
            entry['main'] = value['MSQRoulette']
            entry['mentor'] = value['MentorRoulette']
            entry['normalraid'] = value['NormalRaidRoulette']
            entry['trial'] = value['TrialRoulette']
            return entry
    return entry


def getBeforeAndAfterContentEntries(orderedContent, entry):
    _previous = None
    _next = None
    _type = orderedContent[entry['instanceType']]
    _typeKeys = list(_type)
    for i, k in enumerate(_type):
        if _type[k].endswith(entry['slug']):
            if i - 1 >= 0:
                try:
                    _previous = _type[_typeKeys[i - 1]]
                except:
                    pass
            try:
                _next = _type[_typeKeys[i + 1]]
            except:
                pass
            return _previous, _next
    return None, None


def seperate_data_into_array(tag, entry):
    if entry[tag]:
        entry[tag] = entry[tag].strip("'[").strip("]'").strip("\"[").strip("]\"").strip("[").strip("]").replace("\", \"", "', '").replace("\",\"", "', '").split("', '")
        entry[tag] = [b for b in entry[tag]]


def getEntryData(sheet, max_column, i, elements, orderedContent):
    entry = get_data_from_xlsx(sheet, max_column, i, elements)
    entry = clean_entries_from_single_quotes(entry)
    entry = workOnQuests(entry, entry["quest_id"])
    entry = getEntriesForRouletts(entry)
    for lang in LANGUAGES:
        entry[f"title_{lang}"] = getContentName(entry["title"], lang, entry["difficulty"], entry["instanceType"])
    _previous, _next = getBeforeAndAfterContentEntries(orderedContent, entry)
    # remove time from excel datetime
    entry["date"] = str(entry["date"]).replace(" 00:00:00", "").replace("-", ".")
    entry["prev_content"] = _previous
    entry["next_content"] = _next
    entry["line_index"] = i
    seperate_data_into_array("bosse", entry)
    seperate_data_into_array("adds", entry)
    seperate_data_into_array("tags", entry)
    return entry


def try_to_create_file(filename):
    if not os.path.exists(os.path.dirname(filename)):
        try:
            os.makedirs(os.path.dirname(filename))
        except OSError as exc:  # Guard against race condition
            if exc.errno != errno.EEXIST:
                raise


def addMusic(header_data, music):
    header_data += "music:\n"
    for m in music:
        header_data += f"    - name: \"{m}\"\n"
        _id = getOrchestrionIDByName(m)
        if _id:
            header_data += f"      id: \"{_id}\"\n"
    return header_data


def addContentZoneIdToHeader(header_data, contentzoneid, entry):
    global contentfinderconditionX
    cmt = None
    if not contentzoneid == "":
        header_data += 'contentzoneids:\n'
        for zone in contentzoneid:
            header_data += '  - id: ' + zone + '\n'
    for key, value in contentfinderconditionX.items():
        if value['Name'] == entry['title_de']:
            cmt = value['ContentMemberType']
            if not "InstanceContent" in value['Content']:
                continue
            contentid = value['Content'].replace("InstanceContent#", "")
            if not contentid:
                continue
            _id = "8003" + str(hex(int(contentid))[2:]).rjust(4, '0').upper()
            if "contentzoneids:" not in header_data:
                header_data += 'contentzoneids:\n'

            if _id not in header_data:
                # if _id not in contentzoneid:
                header_data += '  - id: ' + _id + '\n'
    return header_data, cmt


def addGroupCollections(cmt, entry):
    global contentmembertype
    header_data = ""
    skip_lookoup = False
    if not cmt:
        if "Traumprüfung" in entry['title'] or "Dalriada" in entry['title'] or "Castrum Lacus Litore" in entry['title']:
            cmt_entry = 8
            healerp = 2
            tankp = 2
            meleep = 2
            rangep = 2
            skip_lookoup = true
        else:
            print("Could not find GroupCollection for: " + entry['title'])
            return header_data

    if not skip_lookoup:
        wanted_id = cmt.split("#")[1]
        cmt_entry = contentmembertype[f"{wanted_id}"]
        healerp = cmt_entry['HealersPerParty']
        tankp = cmt_entry['TanksPerParty']
        meleep = cmt_entry['MeleesPerParty']
        rangep = cmt_entry['RangedPerParty']

    if int(healerp) + int(tankp) + int(meleep) + int(rangep) > 0:
        header_data += "group:\n"
        if not healerp == "0":
            header_data += f'    healer: "{healerp}"\n'
        if not tankp == "0":
            header_data += f'    tank: "{tankp}"\n'
        if not meleep == "0":
            header_data += f'    melee: "{meleep}"\n'
        if not rangep == "0":
            header_data += f'    range: "{rangep}"\n'
    return header_data


def replaceSlug(text):
    return str(text).replace("_", "-").replace(".", "-").replace(",", "").replace("'", "").replace("ä", "ae").replace("ö", "oe").replace("ü", "ue").replace("Ä", "Ae").replace("Ö", "Oe").replace("Ü", "Ue").replace("ß", "ss")


def getImage(image):
    image = image.replace(".tex", "_hr1.png\"")
    image = image.replace("ui/icon/", "")
    return image


def checkVariable(element, name):
    if element[name] and not element[name] == "###":
        return True
    return False


def get_order_id(entry):
    patch = "0000"
    plvl = "00"
    sortid = "0000"
    if "." in entry["patchNumber"]:
        t_minor = 0
        major, minor = entry["patchNumber"].split(".")
        if str(minor).endswith("a") or str(minor).endswith("b"):
            t_minor = int(minor[:-1])
            t_minor = t_minor * 10 if t_minor < 10 else t_minor
            t_minor = t_minor * 10 if t_minor < 100 else t_minor
            if str(minor).endswith("a"):
                minor = t_minor + 1
            elif str(minor).endswith("b"):
                minor = t_minor + 2
        else:
            minor = f"{minor}0" if len(minor) < 2 else minor
            minor = f"{minor}0" if len(minor) < 3 else minor
        patch = f"{major}{minor}"
    if entry["plvl"]:
        plvl = int(entry["plvl"]) * 10 if len(entry["plvl"]) < 2 else entry["plvl"]
    if entry["sortid"]:
        sortid = f"0{entry['sortid']}" if len(entry["sortid"]) < 1 else entry["sortid"]
        sortid = f"0{sortid}" if len(sortid) < 2 else sortid
        sortid = f"0{sortid}" if len(sortid) < 3 else sortid
        sortid = f"0{sortid}" if len(sortid) < 4 else sortid
    return f"{patch}{plvl}{sortid}"


def addEntries(header_data, entry, field, get_data_function):
    if checkVariable(entry, field):
        header_data += '  - name: "' + entry[field] + '"\n'
        mount_id = get_data_function(entry[field])
        if mount_id:
            header_data += '    id: "' + mount_id + '"\n'
    return header_data


def getMountIDByName(name):
    for _id, mount in mounts.items():
        if mount['Singular_de'] == name:
            return _id
    return None


def getMinionIDByName(name):
    for _id, minion in minions.items():
        if minion['Singular_de'] == name:
            return _id
    return None


def getOrchestrionIDByName(name):
    for _id, orchestrion in orchestrions.items():
        if orchestrion['Name_de'].lower() == name.lower():
            return _id
    return None


def getTTCardIDByName(name):
    for id, ttcard in ttcards.items():
        if ttcard['Name_de'] == name:
            return id.split(".0")[0]
    return None


def get_video_url(url):
    if url.startswith("https"):
        return url
    return "https://www.youtube.com/watch?v={}".format(url)



def get_territorytype_from_mapid(mapid):
    for key, tt_type in territorytype.items():
        if tt_type["TerritoryType"].lower() == mapid.lower():
            return tt_type
    print_color_red(f"Could not find territorytype for {mapid}")
    return ""


def writeTags(header_data, entry, tt_type_name):
    # write tags per expansion
    if entry["categories"] == "arr":
        header_data += "  - term: \"A Realm Reborn\"\n"
        header_data += "  - term: \"ARR\"\n"
    elif entry["categories"] == "hw":
        header_data += "  - term: \"Heavensward\"\n"
        header_data += "  - term: \"HW\"\n"
    elif entry["categories"] == "sb":
        header_data += "  - term: \"Stormblood\"\n"
        header_data += "  - term: \"SB\"\n"
    elif entry["categories"] == "shb":
        header_data += "  - term: \"Shadowbringers\"\n"
        header_data += "  - term: \"ShB\"\n"
    elif entry["categories"] == "ew":
        header_data += "  - term: \"Endwalker\"\n"
        header_data += "  - term: \"EW\"\n"
    else:
        pass

    if not tt_type_name == "":
        for lang in LANGUAGES:
            header_data += "  - term: \"" + tt_type_name["Name_" + lang] + "\"\n"

    for lang in LANGUAGES:
        header_data += "  - term: \"" + entry[f"title_{lang}"] + "\"\n"

    # write rest of the tags
    header_data += "  - term: \"" + entry["difficulty"] + "\"\n"
    header_data += "  - term: \"" + entry["patchNumber"] + "!\"\n"
    header_data += "  - term: \"" + entry["patchName"] + "\"\n"
    if not entry.get("quest", "") == "":
        header_data += "  - term: \"" + entry["quest"] + "\"\n"
    if checkVariable(entry, "mount1") or checkVariable(entry, "mount2"):
        header_data += "  - term: \"mounts\"\n"
        header_data += "  - term: \"Reittier\"\n"
    if checkVariable(entry, "minion1") or checkVariable(entry, "minion2") or checkVariable(entry, "minion3"):
        header_data += "  - term: \"minions\"\n"
        header_data += "  - term: \"Begleiter\"\n"
    if checkVariable(entry, "tt_card1") or checkVariable(entry, "tt_card2"):
        header_data += "  - term: \"tt_cards\"\n"
        header_data += "  - term: \"Triple Triad Karte\"\n"
    if checkVariable(entry, "gearset_loot"):
        for gset in entry["gearset_loot"].split(","):
            header_data += "  - term: \"" + gset + "\"\n"
    if checkVariable(entry, "orchestrion") or checkVariable(entry, "orchestrion2") or checkVariable(entry, "orchestrion3") or checkVariable(entry, "orchestrion4") or checkVariable(entry, "orchestrion5"):
        header_data += "  - term: \"orchestrion\"\n"
        header_data += "  - term: \"Notenrolle\"\n"
    if checkVariable(entry, "orchestrion_material1") or checkVariable(entry, "orchestrion_material2") or checkVariable(entry, "orchestrion_material3"):
        header_data += "  - term: \"orchestrion_material\"\n"
    if entry["instanceType"] == "trial":
        header_data += "  - term: \"Prüfung\"\n"
        header_data += "  - term: \"Trial\"\n"
        header_data += "  - term: \"Primae\"\n"
        header_data += "  - term: \"Primal\"\n"
    header_data += "  - term: \"" + entry["instanceType"] + "\"\n"

    found_roulette = False
    if entry.get("allianceraid", None) == "True":
        header_data += "  - term: \"allianceraid\"\n"
        found_roulette = True
    if entry.get("frontier", None) == "True":
        header_data += "  - term: \"frontier\"\n"
        found_roulette = True
    if entry.get("expert", None) == "True":
        header_data += "  - term: \"expert\"\n"
        found_roulette = True
    if entry.get("guildhest", None) == "True":
        header_data += "  - term: \"guildhest\"\n"
        found_roulette = True
    if entry.get("level50_60_70", None) == "True":
        header_data += "  - term: \"level50_60_70\"\n"
        found_roulette = True
    if entry.get("level80", None) == "True":
        header_data += "  - term: \"level80\"\n"
        found_roulette = True
    if entry.get("leveling", None) == "True":
        header_data += "  - term: \"leveling\"\n"
        found_roulette = True
    if entry.get("main", None) == "True":
        header_data += "  - term: \"main\"\n"
        found_roulette = True
    if entry.get("mentor", None) == "True":
        header_data += "  - term: \"mentor\"\n"
        found_roulette = True
    if entry.get("normalraid", None) == "True":
        header_data += "  - term: \"normalraid\"\n"
        found_roulette = True
    if entry.get("trial", None) == "True":
        header_data += "  - term: \"trial\"\n"
        found_roulette = True
    if found_roulette:
        header_data += "  - term: \"Zufallsinhalt\"\n"
        header_data += "  - term: \"roulette\"\n"

    if not entry["bosse"] == ['']:
        for b in entry["bosse"]:
            if b != "Unknown_":
                header_data += "  - term: \"" + b + "\"\n"
    if not entry["tags"] == ['']:
        for t in entry["tags"]:
            if t != "Unknown_":
                header_data += "  - term: \"" + t + "\"\n"
    return header_data


def rewrite_content_even_if_exists(entry, old_wip):
    header_data = ""
    tt_type_name = get_territorytype_from_mapid(entry["mapid"])
    if old_wip in ["True", "False"]:
        header_data += 'wip: "' + str(old_wip).title() + '"\n'
    else:
        header_data += 'wip: "True"\n'
    #header_data += 'title: "' + entry["title"] + '"\n'
    header_data += 'title:\n'
    for lang in LANGUAGES:
        tmp = entry[f"title_{lang}"].replace(f' ({entry["difficulty"].lower()})', "")
        tmp = tmp.replace(f' ({entry["difficulty"].title()})', "")
        tmp = tmp.replace(f'Traumprüfung - ', "")
        header_data += f'  {lang}: "' + tmp + '"\n'
    header_data += 'layout: guide_post\n'
    header_data += 'page_type: guide\n'
    header_data += f'excel_line: \"{entry["line_index"]}\"\n'
    header_data += 'categories: "' + entry["categories"] + '"\n'
    header_data += 'patchNumber: "' + entry["patchNumber"].replace("'", "") + '"\n'
    if patchversions.get(entry["patchNumber"], None):
        header_data += 'patchLink: "' + patchversions[entry["patchNumber"]]['link_to_patch'] + '"\n'
    header_data += 'difficulty: "' + entry["difficulty"] + '"\n'
    header_data += 'instanceType: "' + entry["instanceType"] + '"\n'
    header_data += 'date: "' + entry["date"] + '"\n'
    header_data += 'slug: "' + replaceSlug(entry["slug"]) + '"\n'
    if entry["prev_content"]:
        header_data += 'previous_slug: "' + replaceSlug(entry["prev_content"]) + '"\n'
    if entry["next_content"]:
        header_data += 'next_slug: "' + replaceSlug(entry["next_content"]) + '"\n'
    if entry["image"]:
        header_data += 'image:\n'
        header_data += '  - url: \"/' + getImage(entry["image"]) + '\n'
        #header_data += '    url: \"/' + getImage(entry["image"]) + '\n'
    header_data += 'terms:\n'
    header_data = writeTags(header_data, entry, tt_type_name)
    header_data += 'patchName: "' + entry["patchName"] + '"\n'
    if entry.get("mapid", None):
        header_data += 'mapid: "' + entry["mapid"] + '"\n'
    if not tt_type_name == "":
        header_data += 'contentname: "' + tt_type_name["Name_de"] + '"\n'
    header_data += 'sortid: ' + entry["sortid"] + '\n'
    header_data += 'plvl: ' + entry["plvl"] + '\n'
    header_data += 'plvl_sync: ' + entry["plvl_sync"] + '\n'
    header_data += 'ilvl: ' + entry["ilvl"] + '\n'
    header_data += 'ilvl_sync: ' + entry["ilvl_sync"] + '\n'
    if not entry["quest"] == "":
        header_data += 'quest: "' + entry["quest"] + '"\n'
    if not entry["quest_location"] == "":
        header_data += 'quest_location: "' + entry["quest_location"] + '"\n'
    if not entry["quest_npc"] == "":
        header_data += 'quest_npc: "' + entry["quest_npc"] + '"\n'
    header_data += 'order: ' + get_order_id(entry) + '\n'
    # mounts
    if checkVariable(entry, "mount1") or checkVariable(entry, "mount2"):
        header_data += 'mount:\n'
        header_data = addEntries(header_data, entry, "mount1", getMountIDByName)
        header_data = addEntries(header_data, entry, "mount2", getMountIDByName)
    # minions
    if checkVariable(entry, "minion1") or checkVariable(entry, "minion2") or checkVariable(entry, "minion3"):
        header_data += 'minion:\n'
        header_data = addEntries(header_data, entry, "minion1", getMinionIDByName)
        header_data = addEntries(header_data, entry, "minion2", getMinionIDByName)
        header_data = addEntries(header_data, entry, "minion3", getMinionIDByName)
    # gearset_loot
    if checkVariable(entry, "gearset_loot"):
        header_data += 'gearset_loot:\n'
        for gset in entry["gearset_loot"] .split(","):
            header_data += '  - gsetname: "' + gset + '"\n'
    # tt_cards
    if checkVariable(entry, "tt_card1") or checkVariable(entry, "tt_card2"):
        header_data += 'tt_card:\n'
        header_data = addEntries(header_data, entry, "tt_card1", getTTCardIDByName)
        header_data = addEntries(header_data, entry, "tt_card2", getTTCardIDByName)
    # orchestrion
    if checkVariable(entry, "orchestrion") or checkVariable(entry, "orchestrion2") or checkVariable(entry, "orchestrion3") or checkVariable(entry, "orchestrion4") or checkVariable(entry, "orchestrion5"):
        header_data += 'orchestrion:\n'
        header_data = addEntries(header_data, entry, "orchestrion", getOrchestrionIDByName)
        header_data = addEntries(header_data, entry, "orchestrion2", getOrchestrionIDByName)
        header_data = addEntries(header_data, entry, "orchestrion3", getOrchestrionIDByName)
        header_data = addEntries(header_data, entry, "orchestrion4", getOrchestrionIDByName)
        header_data = addEntries(header_data, entry, "orchestrion5", getOrchestrionIDByName)
    # orchestrion material
    if checkVariable(entry, "orchestrion_material1") or checkVariable(entry, "orchestrion_material2") or checkVariable(entry, "orchestrion_material3"):
        header_data += 'orchestrion_material:\n'
        if checkVariable(entry, "orchestrion_material1"):
            header_data += '  - name: "' + entry["orchestrion_material1"] + '"\n'
        if checkVariable(entry, "orchestrion_material2"):
            header_data += '  - name: "' + entry["orchestrion_material2"] + '"\n'
        if checkVariable(entry, "orchestrion_material3"):
            header_data += '  - name: "' + entry["orchestrion_material3"] + '"\n'
    # rouletts
    if entry.get("expert", None):
        header_data += 'rouletts:\n'
        if entry["allianceraid"]:
            header_data += '    allianceraid: ' + entry["allianceraid"] + "\n"
        if entry["frontier"]:
            header_data += '    frontier: ' + entry["frontier"] + "\n"
        if entry["expert"]:
            header_data += '    expert: ' + entry["expert"] + "\n"
        if entry["guildhest"]:
            header_data += '    guildhest: ' + entry["guildhest"] + "\n"
        if entry["level50_60_70"]:
            header_data += '    level50_60_70: ' + entry["level50_60_70"] + "\n"
        if entry["level80"]:
            header_data += '    level80: ' + entry["level80"] + "\n"
        if entry["leveling"]:
            header_data += '    leveling: ' + entry["leveling"] + "\n"
        if entry["main"]:
            header_data += '    main: ' + entry["main"] + "\n"
        if entry["mentor"]:
            header_data += '    mentor: ' + entry["mentor"] + "\n"
        if entry["normalraid"]:
            header_data += '    normalraid: ' + entry["normalraid"] + "\n"
        if entry["trial"]:
            header_data += '    trial: ' + entry["trial"] + "\n"
    # links:
    if checkVariable(entry, "teamcraftlink") or checkVariable(entry, "garlandtoolslink") or checkVariable(entry, "gamerescapelink"):
        header_data += 'links:\n'
        if checkVariable(entry, "teamcraftlink"):
            header_data += f'    teamcraftlink: "' + entry["teamcraftlink"] + '"\n'
        if checkVariable(entry, "garlandtoolslink"):
            header_data += f'    garlandtoolslink: "' + entry["garlandtoolslink"] + '"\n'
        if checkVariable(entry, "gamerescapelink"):
            header_data += f'    gamerescapelink: "' + entry["gamerescapelink"] + '"\n'
    # videos
    if checkVariable(entry, "mtqvid1"):
        header_data += 'mtq_vid1: "' + get_video_url(entry["mtqvid1"]) + '"\n'
    if checkVariable(entry, "mtqvid2"):
        header_data += 'mtq_vid2: "' + get_video_url(entry["mtqvid2"]) + '"\n'
    if checkVariable(entry, "mrhvid1"):
        header_data += 'mrh_vid1: "' + get_video_url(entry["mrhvid1"]) + '"\n'
    if checkVariable(entry, "mrhvid2"):
        header_data += 'mrh_vid2: "' + get_video_url(entry["mrhvid2"]) + '"\n'
    return header_data, entry


def addHeader(entry, old_data, music, contentzoneid):
    header_data, entry = rewrite_content_even_if_exists(entry, old_data.get('wip', False))
    header_data, cmt = addContentZoneIdToHeader(header_data, contentzoneid, entry)
    header_data += addGroupCollections(cmt, entry)
    if music:
        header_data = addMusic(header_data, music)
    return header_data


def writeFileIfNoDifferent(filename, filedata):
    try:
        with open(filename, "r", encoding="utf8") as f:
            x_data = f.read()
    except:
        x_data = None

    if not filedata == x_data:
        with open(filename, "w", encoding="utf8") as fi:
            fi.write(filedata)
        print(f"Wrote new data to file {filename}")


def cleanup_logdata(logdata_instance_content):
    try:
        del logdata_instance_content["combatants"]
    except Exception:
        pass
    try:
        del logdata_instance_content["zone"]
    except Exception:
        pass
    try:
        del logdata_instance_content["contentzoneid"]
    except Exception:
        pass
    music = None
    try:
        music = logdata_instance_content["music"]
        del logdata_instance_content["music"]
    except Exception:
        pass
    for enemy_name, enemy in logdata_instance_content.items():
        # try: del enemy["status"]
        # except Exception: pass
        try:
            del enemy["tether"]
        except Exception:
            pass
        try:
            del enemy["headmarker"]
        except Exception:
            pass
    new_lic = {}
    for k, v in logdata_instance_content.items():
        # if not k == "":
        new_lic[k] = v
    return new_lic, music


def getDataFromLogfile(entry):
    logdata_instance_content = None
    music = None
    contentzoneid = ""
    # get correct title capitalization to read data from logdata
    title = uglyContentNameFix(entry["title_de"].title(), entry["instanceType"], entry["difficulty"])
    # get the latest data from logdata
    if not entry["title_de"] == "" and logdata_lower.get(entry["title_de"].lower()):
        try:
            logdata_instance_content = dict(logdata[getContentName(title, lang="de")])
        except Exception:
            logdata_instance_content = dict(logdata[title])
        if logdata_instance_content.get('contentzoneid', None):
            contentzoneid = logdata_instance_content['contentzoneid']
        logdata_instance_content, music = cleanup_logdata(logdata_instance_content)
    return logdata_instance_content, music, contentzoneid


def write_content_to_file(entry, filename, old_data):
    logdata_instance_content, music, contentzoneid = getDataFromLogfile(entry)
    filedata = '---\n'
    filedata += addHeader(entry, old_data, music, contentzoneid)
    filedata += '---'
    filedata += '\n'
    writeFileIfNoDifferent(filename, filedata)


def run(sheet, max_row, max_column, elements, orderedContent):
    # for every row do:
    for i in range(2, max_row):
        try:
            # comment the 2 line out to filter fo a specific line, numbering starts with 1 like it is in excel
            #if i not in [231]:
            #    continue
            entry = getEntryData(sheet, max_column, i, elements, orderedContent)
            logger.info(pretty_json(entry))
            # if the done collumn is not prefilled
            if entry["exclude"] == "end":
                print("END FLAG WAS FOUND!")
                sys.exit(0)
            if not (entry["exclude"] or entry["done"]):
                logger.debug(pretty_json(entry))
                filename = f"ffxiv_content/{entry['categories']}_new/{entry['instanceType']}/{entry['date'].replace('.', '-')}--{entry['patchNumber']}--{entry['sortid'].zfill(5)}--{entry['slug'].replace(',', '')}.md"
                existing_filename = f"ffxiv_content/{entry['categories']}/{entry['instanceType']}/{entry['date'].replace('.', '-')}--{entry['patchNumber']}--{entry['sortid'].zfill(5)}--{entry['slug'].replace(',', '')}.md"
                old_data = {}#get_old_content_if_file_is_found(existing_filename)
                # if old file was found, replace filename to save
                if not old_data == {}:
                    filename = existing_filename
                    # logger.info(pretty_json(old_data))
                try_to_create_file(filename)
                write_content_to_file(entry, filename, old_data)
        except Exception as e:
            logger.critical(f"Error when handeling '{filename}' with line id '{i}' ({e})")
            traceback.print_exception(*sys.exc_info())


if __name__ == "__main__":
    sheet, max_row, max_column = read_xlsx_file()
    # change into _posts dir
    os.chdir("./_posts")
    # first run to create all files
    orderedContent = getPrevAndNextContentOrder(sheet, XLSXELEMENTS, max_row)
    logger.debug(orderedContent)
    run(sheet, max_row, max_column, XLSXELEMENTS, orderedContent)
    # csgf needs also to run from posts dir
    #csgf.run()
    # move back to DEVPOCKETGUIDE dir
    #os.chdir("..")
    #gl.run()
